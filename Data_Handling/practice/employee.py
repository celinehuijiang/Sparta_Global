# reading the file
import openpyxl

workbook = openpyxl.load_workbook('Employee Sample Data.xlsx')
sheet = workbook['Data']
print(sheet)

# getting the list of row
for row in sheet.iter_rows(values_only=True):
    rows = row
    # print(row)

# checking for any duplicates values 
def has_duplicates(lst):
    seen = set()
    for item in lst:
        if item in seen:
            return True
        seen.add(item)
    return False

# checking for any empty columns 
def find_empty_columns(arg_1):
    empty_columns = []
    header = arg_1[0] 
    
    empty_columns = []
    for column in sheet.iter_cols(values_only=True):
        column_data = column[1:]  # Exclude the header row
        if all(value == '' for value in column_data):
            empty_columns.append(column[0])
    
    return empty_columns

find_empty_columns(rows)